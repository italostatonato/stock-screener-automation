import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import DateAxis

logger = logging.getLogger(__name__)

# ── Definição de colunas por tipo de formatação ────────────────────────────────

FII_MONEY_COLS   = ["PREÇO ATUAL (R$)", "LIQUIDEZ DIÁRIA (R$)", "PATRIMÔNIO LÍQUIDO", "VPA"]
FII_PCT_COLS     = [
    "DIVIDEND YIELD", "DY (3M) ACUMULADO", "DY (6M) ACUMULADO", "DY (12M) ACUMULADO",
    "DY (3M) MÉDIA", "DY (6M) MÉDIA", "DY (12M) MÉDIA", "DY ANO",
    "DY PATRIMONIAL", "VARIAÇÃO PREÇO", "RENTAB. PERÍODO", "RENTAB. ACUMULADA",
    "VARIAÇÃO PATRIMONIAL", "RENTAB. PATR. PERÍODO", "RENTAB. PATR. ACUMULADA",
]
FII_NUM_COLS     = ["P/VP", "P/VPA", "ÚLTIMO DIVIDENDO", "VOLATILIDADE"]

ACOES_MONEY_COLS = ["Preço", "Volume Diário Médio (3 meses)", "Market Cap Empresa"]
ACOES_PCT_COLS   = [
    "ROTanC", "ROInvC", "RPL", "ROA", "Margem Líquida", "Margem Bruta", "Margem EBIT",
    "Dividend Yield", "Alavancagem Financeira", "Passivo/Patrimônio Líquido",
]
ACOES_NUM_COLS   = [
    "Preço/Lucro", "Preço/VPA", "Preço/Receita Líquida", "Preço/FCO", "Preço/FCF",
    "Preço/EBIT", "Preço/NCAV", "Preço/Ativo Total", "Preço/Capital Giro",
    "EV/EBIT", "EV/EBITDA", "EV/Receita Líquida", "EV/FCF", "EV/FCO", "EV/Ativo Total",
    "Giro do Ativo Inicial",
]

FMT_MONEY = 'R$ #.##0,00'
FMT_PCT   = '0,00%'
FMT_NUM   = '0,00'

COLOR_HEADER      = "1F4E79"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ALT_ROW     = "EBF3FB"
COLOR_OK           = "C6EFCE"   # verde claro — aprovado
COLOR_RANK         = "FFE699"   # amarelo — rank final
COLOR_ELIMINADO    = "F2F2F2"   # cinza claro — eliminado


def _style_sheet(ws, money_cols=None, pct_cols=None, num_cols=None, highlight_status=False):
    money_cols = money_cols or []
    pct_cols   = pct_cols or []
    num_cols   = num_cols or []

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
    alt_fill    = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_idx = {cell.value: cell.column for cell in ws[1]}

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    status_col_letter = None
    if highlight_status and "Status" in col_idx:
        status_col_letter = get_column_letter(col_idx["Status"])

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()

        if status_col_letter:
            status_val = ws[f"{status_col_letter}{row_idx}"].value or ""
            if status_val.startswith("Rank #"):
                fill = PatternFill("solid", fgColor=COLOR_RANK)
            elif status_val.startswith("Aprovado"):
                fill = PatternFill("solid", fgColor=COLOR_OK)
            elif status_val.startswith("Eliminado"):
                fill = PatternFill("solid", fgColor=COLOR_ELIMINADO)

        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    for col_name, fmt in (
        [(c, FMT_MONEY) for c in money_cols] +
        [(c, FMT_PCT)   for c in pct_cols]   +
        [(c, FMT_NUM)   for c in num_cols]
    ):
        if col_name in col_idx:
            col_letter = get_column_letter(col_idx[col_name])
            for cell in ws[col_letter][1:]:
                cell.number_format = fmt

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"


def _add_premissas(wb, cfg: dict, data_hoje: str, n_fiis: int, n_acoes: int):
    if "Premissas" in wb.sheetnames:
        del wb["Premissas"]
    ws = wb.create_sheet("Premissas")

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
    bold   = Font(bold=True, size=10)
    normal = Font(size=10)
    italic = Font(italic=True, size=9, color="666666")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    f = cfg["filters"]
    a = cfg["filters"]["acoes"]

    rows = [
        ("STOCK SCREENER — PREMISSAS E METODOLOGIA", None),
        (None, None),
        ("Última atualização", data_hoje),
        ("FIIs no Top N",  str(n_fiis)),
        ("Ações no Top N", str(n_acoes)),
        (None, None),
        ("── FONTES DE DADOS ──", None),
        ("FIIs",      "fundsexplorer.com.br/ranking"),
        ("Ações BR",  "investsite.com.br/seleciona_acoes.php"),
        ("Indicadores de mercado", "Banco Central do Brasil (SGS) + AwesomeAPI"),
        (None, None),
        ("── METODOLOGIA: FILTRO EM DUAS CAMADAS ──", None),
        (None, "1) Filtros fixos eliminam ativos com problemas absolutos (dados ausentes, "
               "indicadores fora de qualquer faixa razoável, liquidez/patrimônio mínimos)."),
        (None, "2) Filtros por quartil selecionam os melhores 25% do universo restante "
               "em cada indicador-chave — o corte se adapta automaticamente às condições "
               "do mercado no dia, garantindo que sempre haja candidatos qualificados."),
        (None, "Ativos que passam em todos os filtros são ordenados e os Top N entram "
               "na aba de Recomendações. Os demais aparecem nas abas de Base Completa "
               "com o motivo exato da eliminação, para auditoria e transparência total."),
        (None, None),
        ("── FILTROS FIXOS: FIIs ──", None),
        ("Dividend Yield mínimo",  f"{f['dy_min']*100:.2f}%"),
        ("Liquidez diária mínima", f"R$ {f['liquidez_min']:,.0f}".replace(",", ".")),
        ("Patrimônio mínimo",      f"R$ {f['patrimonio_min']:,.0f}".replace(",", ".")),
        (None, None),
        ("── FILTROS POR QUARTIL: FIIs (calculados a cada execução) ──", None),
        ("P/VP",                "≤ Q25 do universo elegível (mais barato)"),
        ("Dividend Yield",      "≥ Q75 do universo elegível (maior renda)"),
        ("Liquidez diária",     "≥ Q75 do universo elegível (mais líquido)"),
        ("Patrimônio líquido",  "≥ Q75 do universo elegível (maior porte)"),
        ("Volatilidade",        "≤ Q75 do universo elegível (menos volátil)"),
        ("Ordenação final",     "Dividend Yield (desc) → P/VP (asc) → Liquidez (desc)"),
        (None, None),
        ("── FILTROS FIXOS: AÇÕES ──", None),
        ("Volume diário mínimo", f"R$ {a['volume_min']:,.0f}".replace(",", ".")),
        ("Market Cap mínimo",    f"R$ {a['market_cap_min']:,.0f}".replace(",", ".")),
        ("Dividend Yield mínimo", f"{a['dy_min']*100:.2f}%"),
        ("Margem Líquida",        "> 0%"),
        ("ROA",                   "> 0%"),
        (None, None),
        ("── FILTROS POR QUARTIL: AÇÕES (calculados a cada execução) ──", None),
        ("Preço/VPA",  "≤ Q25 do universo elegível (mais barato)"),
        ("Preço/Lucro", "≤ Q25 do universo elegível (mais barato)"),
        ("EV/EBIT",     "≤ Q25 do universo elegível (mais barato)"),
        ("EV/EBITDA",   "≤ Q25 do universo elegível (mais barato)"),
        ("Margem Líquida", "≥ Q75 do universo elegível (mais lucrativa)"),
        ("ROA",          "≥ Q75 do universo elegível (mais eficiente)"),
        ("RPL",          "≥ Q75 do universo elegível (maior retorno s/ PL)"),
        ("Dividend Yield", "≥ Q75 do universo elegível (maior renda)"),
        ("Ordenação final", "Preço/VPA (asc) → Dividend Yield (desc)"),
        (None, None),
        ("── COMO LER AS ABAS DE BASE COMPLETA ──", None),
        (None, "Coluna 'Status': mostra 'Rank #N' se o ativo está no Top N final; "
               "'Aprovado (fora do Top N)' se passou em todos os filtros mas não coube "
               "no corte; ou 'Eliminado no filtro fixo/quartil: <motivo>' explicando "
               "exatamente qual critério reprovou o ativo."),
    ]

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70

    for i, (label, value) in enumerate(rows, start=1):
        cell_a = ws.cell(row=i, column=1, value=label)
        cell_b = ws.cell(row=i, column=2, value=value)

        if i == 1:
            cell_a.font = header_font
            cell_a.fill = header_fill
            cell_b.fill = header_fill
        elif label and label.startswith("──"):
            cell_a.font = bold
        elif label is None and value:
            cell_b.font = italic
            cell_b.alignment = Alignment(wrap_text=True, vertical="top")
        elif label:
            cell_a.font = bold
            cell_b.font = normal

        for cell in (cell_a, cell_b):
            cell.border = border
            if not (label is None and value):
                cell.alignment = Alignment(vertical="center")


def _add_base_completa(wb, sheet_name: str, df: pd.DataFrame, money_cols, pct_cols, num_cols):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    df = df.fillna("").astype(str)
    data = [df.columns.tolist()] + df.values.tolist()
    for row in data:
        ws.append(row)

    _style_sheet(ws, money_cols, pct_cols, num_cols, highlight_status=True)


def _add_indicadores(wb, market_data: dict):
    sheet_name = "Indicadores"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
    bold = Font(bold=True, size=10)

    ws["A1"] = "INDICADORES DE MERCADO"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:F1")

    row_cursor = 3

    # ── Câmbio (tabela resumo no topo) ──────────────────────────────────────
    cambio = market_data.get("cambio", {})
    if cambio:
        ws.cell(row=row_cursor, column=1, value="Câmbio e Cripto (cotação atual)").font = bold
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="Par").font = bold
        ws.cell(row=row_cursor, column=2, value="Valor (R$)").font = bold
        ws.cell(row=row_cursor, column=3, value="Variação dia").font = bold
        row_cursor += 1
        cambio_start_row = row_cursor
        for par, dados in cambio.items():
            ws.cell(row=row_cursor, column=1, value=par)
            c_val = ws.cell(row=row_cursor, column=2, value=dados["valor"])
            c_val.number_format = FMT_MONEY
            c_var = ws.cell(row=row_cursor, column=3, value=dados["variacao_pct"] / 100)
            c_var.number_format = FMT_PCT
            row_cursor += 1
        row_cursor += 2

    # ── Séries históricas (IPCA, Selic, IGP-M) ──────────────────────────────
    series_map = {
        "ipca_12m": "IPCA (variação mensal %)",
        "selic":    "Selic Meta (% a.a.)",
        "igpm":     "IGP-M (variação mensal %)",
    }

    chart_anchor_col = 8  # coluna H — onde os gráficos vão aparecer

    for key, label in series_map.items():
        df_serie = market_data.get(key)
        if df_serie is None or df_serie.empty:
            continue

        ws.cell(row=row_cursor, column=1, value=label).font = bold
        row_cursor += 1
        header_row = row_cursor
        ws.cell(row=row_cursor, column=1, value="Data")
        ws.cell(row=row_cursor, column=2, value="Valor (%)")
        row_cursor += 1
        data_start_row = row_cursor

        for _, r in df_serie.iterrows():
            ws.cell(row=row_cursor, column=1, value=r["data"].strftime("%m/%Y"))
            c = ws.cell(row=row_cursor, column=2, value=float(r["valor"]) / 100)
            c.number_format = FMT_PCT
            row_cursor += 1

        data_end_row = row_cursor - 1

        # gráfico de linha
        chart = LineChart()
        chart.title = label
        chart.height = 7
        chart.width = 14
        chart.style = 10

        values_ref = Reference(ws, min_col=2, min_row=header_row, max_row=data_end_row)
        cats_ref   = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        chart.add_data(values_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        anchor_row = (row_cursor - (data_end_row - data_start_row + 3))
        ws.add_chart(chart, f"{get_column_letter(chart_anchor_col)}{anchor_row}")

        row_cursor += 2

    for col_letter in ["A", "B", "C"]:
        ws.column_dimensions[col_letter].width = 18


def _add_graficos(wb, fii_base: pd.DataFrame, acoes_base: pd.DataFrame):
    sheet_name = "Gráficos"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
    ws["A1"] = "VISÃO GERAL — INDICADORES-CHAVE"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:F1")

    row_cursor = 3

    # ── Histograma de DY dos FIIs aprovados vs eliminados ──────────────────
    if "Status" in fii_base.columns and "DIVIDEND YIELD" in fii_base.columns:
        ws.cell(row=row_cursor, column=1, value="Distribuição de Dividend Yield — FIIs (faixas)").font = Font(bold=True)
        row_cursor += 1
        header_row = row_cursor
        ws.cell(row=row_cursor, column=1, value="Faixa DY")
        ws.cell(row=row_cursor, column=2, value="Qtde de FIIs")
        row_cursor += 1
        data_start = row_cursor

        dy_numeric = pd.to_numeric(fii_base["DIVIDEND YIELD"], errors="coerce").dropna()
        bins = [0, 0.02, 0.04, 0.06, 0.08, 0.10, 1.0]
        labels = ["0-2%", "2-4%", "4-6%", "6-8%", "8-10%", "10%+"]
        faixas = pd.cut(dy_numeric, bins=bins, labels=labels, include_lowest=True)
        contagem = faixas.value_counts().reindex(labels, fill_value=0)

        for label, qtd in contagem.items():
            ws.cell(row=row_cursor, column=1, value=str(label))
            ws.cell(row=row_cursor, column=2, value=int(qtd))
            row_cursor += 1
        data_end = row_cursor - 1

        chart = BarChart()
        chart.title = "FIIs por faixa de Dividend Yield"
        chart.height = 8
        chart.width = 15
        values_ref = Reference(ws, min_col=2, min_row=header_row, max_row=data_end)
        cats_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(values_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "E3")

        row_cursor += 3

    # ── Histograma de P/VPA das Ações ───────────────────────────────────────
    if "Status" in acoes_base.columns and "Preço/VPA" in acoes_base.columns:
        ws.cell(row=row_cursor, column=1, value="Distribuição de Preço/VPA — Ações (faixas)").font = Font(bold=True)
        row_cursor += 1
        header_row = row_cursor
        ws.cell(row=row_cursor, column=1, value="Faixa P/VPA")
        ws.cell(row=row_cursor, column=2, value="Qtde de Ações")
        row_cursor += 1
        data_start = row_cursor

        pvpa_numeric = pd.to_numeric(acoes_base["Preço/VPA"], errors="coerce").dropna()
        bins = [0, 0.5, 1.0, 1.5, 2.0, 3.0, 100]
        labels = ["0-0,5", "0,5-1,0", "1,0-1,5", "1,5-2,0", "2,0-3,0", "3,0+"]
        faixas = pd.cut(pvpa_numeric, bins=bins, labels=labels, include_lowest=True)
        contagem = faixas.value_counts().reindex(labels, fill_value=0)

        for label, qtd in contagem.items():
            ws.cell(row=row_cursor, column=1, value=str(label))
            ws.cell(row=row_cursor, column=2, value=int(qtd))
            row_cursor += 1
        data_end = row_cursor - 1

        chart = BarChart()
        chart.title = "Ações por faixa de Preço/VPA"
        chart.height = 8
        chart.width = 15
        values_ref = Reference(ws, min_col=2, min_row=header_row, max_row=data_end)
        cats_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(values_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "E20")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14


def format_workbook(
    snapshot_path: str,
    cfg: dict,
    data_hoje: str,
    n_fiis: int,
    n_acoes: int,
    fii_base: pd.DataFrame,
    acoes_base: pd.DataFrame,
    market_data: dict,
):
    """Carrega o Excel gerado, aplica formatação e adiciona todas as abas extras."""
    logger.info("Aplicando formatação completa ao Excel...")
    wb = load_workbook(snapshot_path)

    # Recomendações (abas já existentes do save_snapshot)
    if "FII" in wb.sheetnames:
        _style_sheet(wb["FII"], FII_MONEY_COLS, FII_PCT_COLS, FII_NUM_COLS)
    if "Ações BR" in wb.sheetnames:
        _style_sheet(wb["Ações BR"], ACOES_MONEY_COLS, ACOES_PCT_COLS, ACOES_NUM_COLS)

    # Bases completas
    _add_base_completa(wb, "FII — Base Completa", fii_base, FII_MONEY_COLS, FII_PCT_COLS, FII_NUM_COLS)
    _add_base_completa(wb, "Ações — Base Completa", acoes_base, ACOES_MONEY_COLS, ACOES_PCT_COLS, ACOES_NUM_COLS)

    # Gráficos
    _add_graficos(wb, fii_base, acoes_base)

    # Indicadores de mercado
    _add_indicadores(wb, market_data)

    # Premissas
    _add_premissas(wb, cfg, data_hoje, n_fiis, n_acoes)

    # ── Reordena abas ────────────────────────────────────────────────────────
    desired_order = [
        "Ações BR", "FII",                          # Recomendações
        "Indicadores",
        "Premissas",
        "Ações — Base Completa", "FII — Base Completa",  # Bases completas
        "Gráficos",
    ]
    current = wb.sheetnames
    new_order = [name for name in desired_order if name in current]
    new_order += [name for name in current if name not in new_order]  # sobras no fim

    for i, name in enumerate(new_order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(snapshot_path)
    logger.info(f"Formatação concluída: {snapshot_path}")