"""Entrega segura do Excel final.

Esta camada é propositalmente isolada do pipeline de ranking/dados.
Se a entrega falhar, o screener continua válido: o erro é registrado como
warning/error de entrega, sem corromper histórico, dashboard ou modelos.
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import platform
import re
import shutil
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_WINDOWS_DRIVE_RE = re.compile(r"^[A-Za-z]:[\\/]")


@dataclass
class DeliveryTargetResult:
    method: str
    status: str
    message: str
    files: List[str] = field(default_factory=list)


@dataclass
class DeliveryResult:
    status: str
    source_file: str
    sha256: Optional[str] = None
    size_bytes: Optional[int] = None
    generated_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))
    targets: List[DeliveryTargetResult] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        data = asdict(self)
        data["targets"] = [asdict(t) for t in self.targets]
        return data


def _is_windows_path(path: str) -> bool:
    return bool(_WINDOWS_DRIVE_RE.match(str(path)))


def _is_usable_local_path(path: str) -> bool:
    """Evita criar `C:/Users/...` falso no runner Linux do GitHub Actions."""
    if _is_windows_path(path) and platform.system().lower() != "windows":
        return False
    return True


def sha256_file(path: str | Path, chunk_size: int = 1024 * 1024) -> str:
    hasher = hashlib.sha256()
    with Path(path).open("rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def validate_excel_file(
    excel_path: str | Path,
    expected_sheets: Optional[Iterable[str]] = None,
) -> Dict[str, Any]:
    """Valida o arquivo Excel antes de entregar.

    Regras bloqueantes: existir, ser .xlsx, tamanho > 0 e abrir com openpyxl.
    Abas esperadas são tratadas como warning configurável, para não quebrar
    o pipeline se a estrutura do workbook evoluir.
    """
    path = Path(excel_path)
    checks: List[Dict[str, str]] = []

    if not path.exists():
        raise FileNotFoundError(f"Excel não encontrado: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Arquivo final precisa ser .xlsx: {path}")

    size = path.stat().st_size
    if size <= 0:
        raise ValueError(f"Excel vazio: {path}")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheetnames = list(workbook.sheetnames)
        workbook.close()
    except Exception as exc:
        raise ValueError(f"Excel inválido ou corrompido: {path}. Erro: {exc}") from exc

    if not sheetnames:
        raise ValueError(f"Excel sem abas: {path}")

    expected = [s for s in (expected_sheets or []) if s]
    missing = sorted(set(expected) - set(sheetnames))
    if missing:
        checks.append({
            "status": "warn",
            "detail": f"Abas esperadas ausentes: {', '.join(missing)}",
        })

    return {
        "status": "ok" if not checks else "warn",
        "path": str(path),
        "size_bytes": size,
        "sheetnames": sheetnames,
        "checks": checks,
    }


def _atomic_copy(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    tmp = dst.with_suffix(dst.suffix + ".tmp")
    if tmp.exists():
        tmp.unlink()
    shutil.copy2(src, tmp)
    os.replace(tmp, dst)


def _delivery_config(cfg: Dict[str, Any]) -> Dict[str, Any]:
    paths = cfg.get("paths", {}) if isinstance(cfg, dict) else {}
    delivery = cfg.get("delivery", {}) if isinstance(cfg, dict) else {}
    excel_cfg = delivery.get("excel", {}) if isinstance(delivery, dict) else {}

    # Compatibilidade com o projeto atual: paths.onedrive_output_dir já existe.
    legacy_path = paths.get("onedrive_output_dir")

    local_copy = excel_cfg.get("local_copy", {}) if isinstance(excel_cfg, dict) else {}
    target_path = local_copy.get("path") or legacy_path

    enabled = excel_cfg.get("enabled")
    if enabled is None:
        enabled = bool(target_path)

    return {
        "enabled": bool(enabled),
        "expected_sheets": excel_cfg.get("expected_sheets", []),
        "keep_dated_copy": bool(local_copy.get("keep_dated_copy", True)),
        "keep_latest_copy": bool(local_copy.get("keep_latest_copy", True)),
        "latest_filename": local_copy.get("latest_filename", "Top20_Ranking_Atual.xlsx"),
        "local_copy_enabled": bool(local_copy.get("enabled", bool(target_path))),
        "local_copy_path": target_path,
        "log_enabled": bool(excel_cfg.get("log_enabled", True)),
        "log_dir": excel_cfg.get("log_dir") or os.path.join(paths.get("data_dir", "data"), "delivery"),
        "explicit_delivery_cfg": bool(excel_cfg),
    }


def _write_delivery_log(result: DeliveryResult, log_dir: str | Path) -> None:
    path = Path(log_dir) / "delivery_log.jsonl"
    path.parent.mkdir(parents=True, exist_ok=True)

    # Não grava caminho completo de destino para evitar expor detalhes locais.
    safe = result.to_dict()
    for target in safe.get("targets", []):
        target["files"] = [Path(f).name for f in target.get("files", [])]

    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(safe, ensure_ascii=False, sort_keys=True) + "\n")


def copy_to_local_onedrive(
    excel_path: str | Path,
    destination_dir: str | Path,
    *,
    keep_dated_copy: bool = True,
    keep_latest_copy: bool = True,
    latest_filename: str = "Top20_Ranking_Atual.xlsx",
) -> DeliveryTargetResult:
    src = Path(excel_path)
    destination = Path(destination_dir)

    if not str(destination_dir).strip():
        return DeliveryTargetResult("local_onedrive", "skipped", "Destino local não configurado.")

    if not _is_usable_local_path(str(destination_dir)):
        return DeliveryTargetResult(
            "local_onedrive",
            "skipped",
            "Destino Windows local ignorado fora do Windows. Use Microsoft Graph para GitHub Actions.",
        )

    copied: List[str] = []
    if keep_dated_copy:
        dated_dst = destination / src.name
        _atomic_copy(src, dated_dst)
        copied.append(str(dated_dst))

    if keep_latest_copy:
        latest_dst = destination / latest_filename
        _atomic_copy(src, latest_dst)
        copied.append(str(latest_dst))

    if not copied:
        return DeliveryTargetResult("local_onedrive", "skipped", "Nenhuma cópia habilitada.")

    return DeliveryTargetResult(
        "local_onedrive",
        "ok",
        f"{len(copied)} arquivo(s) entregue(s) no diretório local sincronizado.",
        copied,
    )


def deliver_excel(
    snapshot_path: str | Path,
    cfg: Dict[str, Any],
    *,
    data_execucao: Optional[str] = None,
) -> DeliveryResult:
    """Entrega o Excel final conforme configuração.

    Hoje implementa a fase 1: cópia local para pasta sincronizada do OneDrive.
    A estrutura já está preparada para receber Microsoft Graph na fase 2.
    """
    delivery_cfg = _delivery_config(cfg)
    source = Path(snapshot_path)

    result = DeliveryResult(status="skipped", source_file=str(source))

    if not delivery_cfg["enabled"]:
        result.targets.append(DeliveryTargetResult("delivery", "skipped", "Entrega desabilitada."))
        return result

    try:
        validation = validate_excel_file(source, delivery_cfg.get("expected_sheets"))
        result.sha256 = sha256_file(source)
        result.size_bytes = int(validation["size_bytes"])
    except Exception as exc:
        result.status = "error"
        result.targets.append(DeliveryTargetResult("validation", "error", str(exc)))
        logger.error("Falha na validação do Excel final: %s", exc)
        if delivery_cfg.get("log_enabled"):
            _write_delivery_log(result, delivery_cfg["log_dir"])
        return result

    target_results: List[DeliveryTargetResult] = []

    if delivery_cfg["local_copy_enabled"]:
        try:
            target_results.append(
                copy_to_local_onedrive(
                    source,
                    delivery_cfg["local_copy_path"],
                    keep_dated_copy=delivery_cfg["keep_dated_copy"],
                    keep_latest_copy=delivery_cfg["keep_latest_copy"],
                    latest_filename=delivery_cfg["latest_filename"],
                )
            )
        except Exception as exc:
            target_results.append(DeliveryTargetResult("local_onedrive", "error", str(exc)))
            logger.error("Falha na entrega local do Excel: %s", exc)

    result.targets.extend(target_results)

    statuses = {t.status for t in result.targets}
    if "ok" in statuses and "error" not in statuses:
        result.status = "ok"
    elif "error" in statuses:
        result.status = "warn" if "ok" in statuses or "skipped" in statuses else "error"
    else:
        result.status = "skipped"

    # Evita criar logs inúteis no GitHub Actions quando apenas pulou path Windows legado.
    should_log = delivery_cfg.get("log_enabled") and (
        result.status != "skipped" or delivery_cfg.get("explicit_delivery_cfg")
    )
    if should_log:
        _write_delivery_log(result, delivery_cfg["log_dir"])

    return result
