from pathlib import Path

from openpyxl import Workbook

from src.delivery import copy_to_local_onedrive, deliver_excel, sha256_file, validate_excel_file


def _make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "ok"
    wb.save(path)


def test_validate_excel_and_hash(tmp_path):
    xlsx = tmp_path / "Top20_Ranking_2026-07-04.xlsx"
    _make_workbook(xlsx)

    validation = validate_excel_file(xlsx)

    assert validation["status"] == "ok"
    assert validation["size_bytes"] > 0
    assert "Resumo" in validation["sheetnames"]
    assert len(sha256_file(xlsx)) == 64


def test_copy_to_local_onedrive_creates_dated_and_latest(tmp_path):
    xlsx = tmp_path / "Top20_Ranking_2026-07-04.xlsx"
    dest = tmp_path / "onedrive"
    _make_workbook(xlsx)

    result = copy_to_local_onedrive(xlsx, dest)

    assert result.status == "ok"
    assert (dest / "Top20_Ranking_2026-07-04.xlsx").exists()
    assert (dest / "Top20_Ranking_Atual.xlsx").exists()


def test_deliver_excel_uses_legacy_onedrive_path(tmp_path):
    xlsx = tmp_path / "Top20_Ranking_2026-07-04.xlsx"
    dest = tmp_path / "onedrive"
    _make_workbook(xlsx)

    cfg = {
        "paths": {
            "data_dir": str(tmp_path / "data"),
            "onedrive_output_dir": str(dest),
        }
    }

    result = deliver_excel(xlsx, cfg)

    assert result.status == "ok"
    assert (dest / "Top20_Ranking_2026-07-04.xlsx").exists()
    assert (dest / "Top20_Ranking_Atual.xlsx").exists()
    assert (tmp_path / "data" / "delivery" / "delivery_log.jsonl").exists()
