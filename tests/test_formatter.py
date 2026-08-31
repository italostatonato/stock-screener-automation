from openpyxl import Workbook

from src.formatter import _add_indicadores, _add_premissas


def test_add_indicadores_aceita_cambio_sem_variacao_diaria():
    wb = Workbook()

    _add_indicadores(
        wb,
        {
            "cambio": {
                "USD/BRL": {
                    "valor": 5.19,
                    "variacao_pct": None,
                }
            }
        },
    )

    ws = wb["Indicadores"]
    assert ws["B5"].value == 5.19
    assert ws["C5"].value is None


def test_premissas_documentam_score_igual_sem_filtros_de_quartil():
    wb = Workbook()
    cfg = {
        "filters": {
            "dy_min": 0.003,
            "liquidez_min": 100_000,
            "patrimonio_min": 50_000_000,
            "acoes": {
                "dy_min": 0.02,
                "volume_min": 500_000,
                "market_cap_min": 200_000_000,
            },
        }
    }

    _add_premissas(wb, cfg, "2026-08-31", 20, 20)

    values = [cell.value for row in wb["Premissas"].iter_rows() for cell in row]
    assert "── SCORE: FIIs — 1/7 CADA ──" in values
    assert "── SCORE: AÇÕES — 1/7 CADA ──" in values
    assert not any("QUARTIL" in str(value) for value in values)
